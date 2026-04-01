from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, date
import io
import openpyxl
from app.database import get_db
from app.models.models import (
    Customer, CashFlow, FixedExpense,
    MonthlySummary, FinancialGoal, RealEstateAnalysis,
    InvestmentStatus, FinancialSnapshot, LedgerTransaction, UploadHistory,
)
from app.schemas.schemas import (
    CustomerCreate, CustomerUpdate, CustomerResponse,
    CashFlowCreate, CashFlowUpdate, CashFlowResponse,
    FixedExpenseCreate, FixedExpenseUpdate, FixedExpenseResponse,
    MonthlySummaryCreate, MonthlySummaryUpdate, MonthlySummaryResponse,
    FinancialGoalCreate, FinancialGoalUpdate, FinancialGoalResponse,
    RealEstateAnalysisCreate, RealEstateAnalysisUpdate, RealEstateAnalysisResponse,
    InvestmentStatusCreate, InvestmentStatusUpdate, InvestmentStatusResponse,
    FinancialSnapshotResponse,
    LedgerTransactionCreate, LedgerTransactionUpdate, LedgerTransactionResponse,
    UploadHistoryResponse,
)

router = APIRouter()


# ── Customer ─────────────────────────────────────────────────
@router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return db.query(Customer).offset(skip).limit(limit).all()

@router.post("/customers", response_model=CustomerResponse)
async def create_customer(data: CustomerCreate, db: Session = Depends(get_db)):
    obj = Customer(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/customers/{customer_id}", response_model=CustomerResponse)
async def update_customer(customer_id: int, data: CustomerUpdate, db: Session = Depends(get_db)):
    obj = db.query(Customer).filter(Customer.id == customer_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="고객 정보를 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    obj = db.query(Customer).filter(Customer.id == customer_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="고객 정보를 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ── CashFlow ──────────────────────────────────────────────────
@router.get("/cash-flows", response_model=List[CashFlowResponse])
async def get_cash_flows(skip: int = 0, limit: int = 1000, db: Session = Depends(get_db)):
    return db.query(CashFlow).offset(skip).limit(limit).all()

@router.post("/cash-flows", response_model=CashFlowResponse)
async def create_cash_flow(data: CashFlowCreate, db: Session = Depends(get_db)):
    obj = CashFlow(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/cash-flows/{cash_flow_id}", response_model=CashFlowResponse)
async def update_cash_flow(cash_flow_id: int, data: CashFlowUpdate, db: Session = Depends(get_db)):
    obj = db.query(CashFlow).filter(CashFlow.id == cash_flow_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="현금흐름 항목을 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/cash-flows/{cash_flow_id}")
async def delete_cash_flow(cash_flow_id: int, db: Session = Depends(get_db)):
    obj = db.query(CashFlow).filter(CashFlow.id == cash_flow_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="현금흐름 항목을 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ── FixedExpense ──────────────────────────────────────────────
@router.get("/fixed-expenses", response_model=List[FixedExpenseResponse])
async def get_fixed_expenses(
    skip: int = 0, limit: int = 1000, db: Session = Depends(get_db)
):
    return db.query(FixedExpense).offset(skip).limit(limit).all()

@router.post("/fixed-expenses", response_model=FixedExpenseResponse)
async def create_fixed_expense(data: FixedExpenseCreate, db: Session = Depends(get_db)):
    obj = FixedExpense(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/fixed-expenses/{expense_id}", response_model=FixedExpenseResponse)
async def update_fixed_expense(expense_id: int, data: FixedExpenseUpdate, db: Session = Depends(get_db)):
    obj = db.query(FixedExpense).filter(FixedExpense.id == expense_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="고정비 항목을 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/fixed-expenses/{expense_id}")
async def delete_fixed_expense(expense_id: int, db: Session = Depends(get_db)):
    obj = db.query(FixedExpense).filter(FixedExpense.id == expense_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="고정비 항목을 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ── MonthlySummary ────────────────────────────────────────────
@router.get("/monthly-summaries", response_model=List[MonthlySummaryResponse])
async def get_monthly_summaries(
    year: Optional[int] = None,
    skip: int = 0, limit: int = 1000,
    db: Session = Depends(get_db),
):
    q = db.query(MonthlySummary)
    if year:
        q = q.filter(MonthlySummary.year == year)
    return q.order_by(MonthlySummary.year, MonthlySummary.month).offset(skip).limit(limit).all()

@router.post("/monthly-summaries", response_model=MonthlySummaryResponse)
async def create_monthly_summary(data: MonthlySummaryCreate, db: Session = Depends(get_db)):
    obj = MonthlySummary(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/monthly-summaries/{summary_id}", response_model=MonthlySummaryResponse)
async def update_monthly_summary(summary_id: int, data: MonthlySummaryUpdate, db: Session = Depends(get_db)):
    obj = db.query(MonthlySummary).filter(MonthlySummary.id == summary_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="월별 결산 항목을 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/monthly-summaries/{summary_id}")
async def delete_monthly_summary(summary_id: int, db: Session = Depends(get_db)):
    obj = db.query(MonthlySummary).filter(MonthlySummary.id == summary_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="월별 결산 항목을 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ── FinancialGoal ─────────────────────────────────────────────
@router.get("/financial-goals", response_model=List[FinancialGoalResponse])
async def get_financial_goals(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return db.query(FinancialGoal).offset(skip).limit(limit).all()

@router.post("/financial-goals", response_model=FinancialGoalResponse)
async def create_financial_goal(data: FinancialGoalCreate, db: Session = Depends(get_db)):
    obj = FinancialGoal(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/financial-goals/{goal_id}", response_model=FinancialGoalResponse)
async def update_financial_goal(goal_id: int, data: FinancialGoalUpdate, db: Session = Depends(get_db)):
    obj = db.query(FinancialGoal).filter(FinancialGoal.id == goal_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="재무 목표를 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/financial-goals/{goal_id}")
async def delete_financial_goal(goal_id: int, db: Session = Depends(get_db)):
    obj = db.query(FinancialGoal).filter(FinancialGoal.id == goal_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="재무 목표를 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ── RealEstateAnalysis ────────────────────────────────────────
@router.get("/real-estate-analyses", response_model=List[RealEstateAnalysisResponse])
async def get_real_estate_analyses(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return db.query(RealEstateAnalysis).offset(skip).limit(limit).all()

@router.post("/real-estate-analyses", response_model=RealEstateAnalysisResponse)
async def create_real_estate_analysis(data: RealEstateAnalysisCreate, db: Session = Depends(get_db)):
    obj = RealEstateAnalysis(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/real-estate-analyses/{analysis_id}", response_model=RealEstateAnalysisResponse)
async def update_real_estate_analysis(analysis_id: int, data: RealEstateAnalysisUpdate, db: Session = Depends(get_db)):
    obj = db.query(RealEstateAnalysis).filter(RealEstateAnalysis.id == analysis_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="부동산 수익분석을 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/real-estate-analyses/{analysis_id}")
async def delete_real_estate_analysis(analysis_id: int, db: Session = Depends(get_db)):
    obj = db.query(RealEstateAnalysis).filter(RealEstateAnalysis.id == analysis_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="부동산 수익분석을 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ── InvestmentStatus ──────────────────────────────────────────
@router.get("/investment-statuses", response_model=List[InvestmentStatusResponse])
async def get_investment_statuses(skip: int = 0, limit: int = 200, db: Session = Depends(get_db)):
    return db.query(InvestmentStatus).order_by(InvestmentStatus.id).offset(skip).limit(limit).all()

@router.post("/investment-statuses", response_model=InvestmentStatusResponse)
async def create_investment_status(data: InvestmentStatusCreate, db: Session = Depends(get_db)):
    obj = InvestmentStatus(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/investment-statuses/{status_id}", response_model=InvestmentStatusResponse)
async def update_investment_status(status_id: int, data: InvestmentStatusUpdate, db: Session = Depends(get_db)):
    obj = db.query(InvestmentStatus).filter(InvestmentStatus.id == status_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="투자 현황 항목을 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/investment-statuses/{status_id}")
async def delete_investment_status(status_id: int, db: Session = Depends(get_db)):
    obj = db.query(InvestmentStatus).filter(InvestmentStatus.id == status_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="투자 현황 항목을 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ── FinancialSnapshot ─────────────────────────────────────────
@router.get("/financial-snapshot", response_model=Optional[FinancialSnapshotResponse])
async def get_financial_snapshot(db: Session = Depends(get_db)):
    return db.query(FinancialSnapshot).order_by(FinancialSnapshot.id.desc()).first()


# ── LedgerTransaction ─────────────────────────────────────────
@router.get("/ledger-transactions", response_model=List[LedgerTransactionResponse])
async def get_ledger_transactions(
    transaction_type: Optional[str] = None,
    category: Optional[str] = None,
    skip: int = 0,
    limit: int = 5000,
    db: Session = Depends(get_db),
):
    q = db.query(LedgerTransaction)
    if transaction_type:
        q = q.filter(LedgerTransaction.transaction_type == transaction_type)
    if category:
        q = q.filter(LedgerTransaction.category == category)
    return q.order_by(LedgerTransaction.transaction_date.desc()).offset(skip).limit(limit).all()

@router.post("/ledger-transactions", response_model=LedgerTransactionResponse)
async def create_ledger_transaction(data: LedgerTransactionCreate, db: Session = Depends(get_db)):
    obj = LedgerTransaction(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj

@router.put("/ledger-transactions/{tx_id}", response_model=LedgerTransactionResponse)
async def update_ledger_transaction(tx_id: int, data: LedgerTransactionUpdate, db: Session = Depends(get_db)):
    obj = db.query(LedgerTransaction).filter(LedgerTransaction.id == tx_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="가계부 내역을 찾을 수 없습니다.")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj

@router.delete("/ledger-transactions/{tx_id}")
async def delete_ledger_transaction(tx_id: int, db: Session = Depends(get_db)):
    obj = db.query(LedgerTransaction).filter(LedgerTransaction.id == tx_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="가계부 내역을 찾을 수 없습니다.")
    db.delete(obj)
    db.commit()
    return {"ok": True}


@router.get("/upload-history", response_model=List[UploadHistoryResponse])
async def get_upload_history(limit: int = 50, db: Session = Depends(get_db)):
    """업로드 이력 목록 (최신순)"""
    return (
        db.query(UploadHistory)
        .order_by(UploadHistory.created_at.desc())
        .limit(limit)
        .all()
    )


@router.post("/import/banksalad-excel")
async def import_banksalad_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """뱅크샐러드 Excel 파일을 파싱하여 고객·현금흐름·월별결산·가계부 내역을 일괄 갱신합니다."""
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="xlsx 또는 xls 파일만 지원합니다.")

    file_bytes = await file.read()
    file_size = len(file_bytes)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 파일을 읽을 수 없습니다.")

    result: dict = {
        "customer": {"updated": 0, "inserted": 0},
        "cash_flow": {"updated": 0, "inserted": 0},
        "monthly_summary": {"updated": 0, "inserted": 0},
        "investment": {"updated": 0, "inserted": 0},
        "financial_snapshot": {"updated": 0, "inserted": 0},
        "ledger": {"inserted": 0, "skipped": 0},
    }

    # ── 기존 오염 데이터 정리 ──────────────────────────────────
    # company/principal 없이 삽입된 투자 레코드 제거 (이전 import 버그로 인한 잉여 레코드)
    db.query(InvestmentStatus).filter(
        InvestmentStatus.company == None,
        InvestmentStatus.principal == None,
        InvestmentStatus.return_rate == None,
    ).delete(synchronize_session=False)
    # 수입·지출·순수익이 모두 0인 월별 결산 레코드 제거 (빈 월 레코드)
    db.query(MonthlySummary).filter(
        MonthlySummary.income == 0,
        MonthlySummary.expense == 0,
        MonthlySummary.net_income == 0,
    ).delete(synchronize_session=False)
    db.query(MonthlySummary).filter(
        MonthlySummary.income == None,
        MonthlySummary.expense == None,
        MonthlySummary.net_income == None,
    ).delete(synchronize_session=False)
    db.commit()

    # ── 뱅샐현황 시트 ─────────────────────────────────────────
    if "뱅샐현황" in wb.sheetnames:
        ws_bs = wb["뱅샐현황"]

        # 행별 데이터 추출 (rows 1~120, cols B~P = cols 2~16)
        rows: dict[int, tuple] = {}
        for ws_row in ws_bs.iter_rows(min_row=1, max_row=120, min_col=2, max_col=16):
            row_idx = ws_row[0].row
            rows[row_idx] = tuple(cell.value for cell in ws_row)

        # 1. 고객 정보 (Row 6: B=이름, C=성별, D=나이, E=신용점수, F=이메일)
        r6 = rows.get(6)
        if r6 and r6[0]:
            name = str(r6[0])
            gender = str(r6[1]) if r6[1] else None
            age = int(r6[2]) if r6[2] is not None else None
            credit_score = int(r6[3]) if r6[3] is not None else None
            email_raw = str(r6[4]) if r6[4] else None
            email = email_raw if email_raw and email_raw != '-' else None

            customer = db.query(Customer).filter(Customer.name == name).first()
            if customer:
                customer.gender = gender
                customer.age = age
                customer.credit_score = credit_score
                if email:
                    customer.email = email
                result["customer"]["updated"] += 1
            else:
                db.add(Customer(name=name, gender=gender, age=age, credit_score=credit_score, email=email))
                result["customer"]["inserted"] += 1
            db.commit()

        # 2. 현금흐름 항목 파싱 — 헤더('항목')를 동적으로 찾아 월 레이블 추출
        #    이후 '월수입 총계' / '월지출 총계' / '순수입 총계' 키워드로 경계 탐지
        #    (Excel 버전마다 항목 수가 달라 행 번호가 변동됨)
        header_row_idx = None
        for ridx in range(9, 20):
            r = rows.get(ridx)
            if r and r[0] == '항목':
                header_row_idx = ridx
                break
        r_header = rows.get(header_row_idx, (None,) * 15) if header_row_idx else (None,) * 15
        month_labels = [str(r_header[i]) for i in range(3, 15) if r_header and r_header[i] is not None]

        # 키워드로 경계 행 찾기
        income_total_row: "int | None" = None
        expense_total_row: "int | None" = None
        net_total_row: "int | None" = None
        for ridx in range(10, 45):
            r = rows.get(ridx)
            if r is None:
                continue
            label = str(r[0]) if r[0] else ''
            if label == '월수입 총계' and income_total_row is None:
                income_total_row = ridx
            elif label == '월지출 총계' and expense_total_row is None:
                expense_total_row = ridx
            elif label == '순수입 총계' and net_total_row is None:
                net_total_row = ridx

        SKIP_CASHFLOW = {'월수입 총계', '월지출 총계', '순수입 총계'}

        def _upsert_cashflow(row_data: tuple, item_type: str) -> None:
            item_name = str(row_data[0]) if row_data[0] else None
            if not item_name or item_name in SKIP_CASHFLOW:
                return
            total = float(row_data[1]) if row_data[1] is not None else None
            monthly_avg = float(row_data[2]) if row_data[2] is not None else None
            monthly_data = {
                month_labels[j]: float(row_data[3 + j])
                for j in range(len(month_labels))
                if 3 + j < len(row_data) and row_data[3 + j] is not None
            }
            # Excel 합계 컬럼이 수식(formula)이어서 0으로 읽히는 경우 monthly_data 합계로 대체
            if not total and monthly_data:
                total = sum(monthly_data.values())
            if not monthly_avg and total and month_labels:
                monthly_avg = total / len(month_labels)
            existing = db.query(CashFlow).filter(CashFlow.item_name == item_name).first()
            if existing:
                existing.item_type = item_type
                existing.total = total
                existing.monthly_average = monthly_avg
                existing.monthly_data = monthly_data
                result["cash_flow"]["updated"] += 1
            else:
                db.add(CashFlow(
                    item_name=item_name, item_type=item_type,
                    total=total, monthly_average=monthly_avg, monthly_data=monthly_data,
                ))
                result["cash_flow"]["inserted"] += 1

        # 헤더 다음 행 ~ 월수입 총계 직전 → 수입 항목
        if header_row_idx and income_total_row:
            for ridx in range(header_row_idx + 1, income_total_row):
                if ridx in rows:
                    _upsert_cashflow(rows[ridx], "수입")
        # 월수입 총계 다음 행 ~ 월지출 총계 직전 → 지출 항목
        if income_total_row and expense_total_row:
            for ridx in range(income_total_row + 1, expense_total_row):
                if ridx in rows:
                    _upsert_cashflow(rows[ridx], "지출")
        db.commit()

        # 3. 월별 결산 upsert — 동적으로 찾은 총계 행 사용
        r_income_total = rows.get(income_total_row) if income_total_row else None
        r_expense_total = rows.get(expense_total_row) if expense_total_row else None
        r_net_total = rows.get(net_total_row) if net_total_row else None

        # 총계 행의 월별 값도 수식 캐시 없으면 0 → 수입/지출 항목 합계로 대체
        def _monthly_sum_from_items(item_type_filter: str, j: int) -> "float | None":
            items = db.query(CashFlow).filter(CashFlow.item_type == item_type_filter).all()
            total_j = sum(
                float(cf.monthly_data.get(month_labels[j], 0))
                for cf in items
                if cf.monthly_data and j < len(month_labels)
            )
            return total_j if total_j else None

        def _safe_float(row_data: "tuple | None", idx: int) -> "float | None":
            if row_data is None:
                return None
            return float(row_data[idx]) if len(row_data) > idx and row_data[idx] is not None else None

        for j, label in enumerate(month_labels):
            try:
                year_val, month_val = int(label[:4]), int(label[5:7])
            except (ValueError, IndexError):
                continue

            income = _safe_float(r_income_total, 3 + j)
            expense = _safe_float(r_expense_total, 3 + j)
            net = _safe_float(r_net_total, 3 + j)

            # 수식 캐시 없어서 0인 경우 항목 합산으로 대체
            if not income:
                income = _monthly_sum_from_items("수입", j)
            if not expense:
                expense = _monthly_sum_from_items("지출", j)
            if income is not None and expense is not None and not net:
                net = income - expense

            # 수입·지출·순수익이 모두 None 또는 0이면 의미 없는 행이므로 건너뜀
            if income is None and expense is None and net is None:
                continue
            if (income or 0) == 0 and (expense or 0) == 0 and (net or 0) == 0:
                continue

            existing_ms = db.query(MonthlySummary).filter(
                MonthlySummary.year == year_val,
                MonthlySummary.month == month_val,
            ).first()
            if existing_ms:
                existing_ms.income = income
                existing_ms.expense = expense
                existing_ms.net_income = net
                result["monthly_summary"]["updated"] += 1
            else:
                db.add(MonthlySummary(year=year_val, month=month_val, income=income, expense=expense, net_income=net))
                result["monthly_summary"]["inserted"] += 1

        db.commit()

        # cumulative_net_income 재계산 (연도별 누적)
        from itertools import groupby as _groupby
        all_sums = db.query(MonthlySummary).order_by(MonthlySummary.year, MonthlySummary.month).all()
        for _year, _group in _groupby(all_sums, key=lambda s: s.year):
            _cumulative = 0.0
            for s in _group:
                if s.net_income is not None:
                    _cumulative += float(s.net_income)
                s.cumulative_net_income = _cumulative
        db.commit()

        # 4. 투자성 자산 (뱅샐현황 rows 73~102, '투자성 자산' 섹션)
        # col B(index 0)=섹션라벨, col C(index 1)=상품명, col E(index 3)=평가금액
        from collections import defaultdict as _defaultdict

        excel_investments: list[tuple[str, float]] = []
        in_inv_section = False
        for ridx in range(70, 120):
            r = rows.get(ridx)
            if r is None:
                continue
            label = r[0]  # col B
            if label == '투자성 자산':
                in_inv_section = True
            elif in_inv_section and label is not None:
                break  # 다음 섹션 진입 → 종료

            if not in_inv_section:
                continue

            product_name = str(r[1]) if r[1] else None
            raw_val = r[3]  # col E
            if not product_name or raw_val is None:
                continue
            try:
                val = float(raw_val)
            except (TypeError, ValueError):
                continue
            if val <= 1:  # 사실상 0인 항목 제외
                continue
            excel_investments.append((product_name, val))

        if excel_investments:
            # DB 레코드를 상품명별로 그룹화 (ID 오름차순 = 마이그레이션 삽입 순서)
            all_inv_db = db.query(InvestmentStatus).order_by(InvestmentStatus.id).all()
            db_by_name: dict[str, list] = _defaultdict(list)
            for inv in all_inv_db:
                db_by_name[inv.product_name].append(inv)

            # 엑셀도 상품명별로 순서대로 그룹화
            excel_by_name: dict[str, list[float]] = _defaultdict(list)
            for pname, val in excel_investments:
                excel_by_name[pname].append(val)

            for pname, excel_vals in excel_by_name.items():
                db_records = db_by_name.get(pname, [])
                for i, val in enumerate(excel_vals):
                    if i < len(db_records):
                        rec = db_records[i]
                        rec.current_value = val
                        # 원금이 있으면 수익률 재계산
                        if rec.principal and float(rec.principal) > 0:
                            rec.return_rate = (val - float(rec.principal)) / float(rec.principal) * 100
                        result["investment"]["updated"] += 1
                    # DB에 없는 상품은 신규 삽입하지 않음 (기존 마이그레이션 데이터만 갱신)

            db.commit()

        # investment_principal / investment_value → 가장 최근 MonthlySummary에 반영
        # investment_value: 엑셀 투자성 자산 섹션의 평가금액 합계
        # investment_principal: InvestmentStatus DB의 원금 합계
        if excel_investments:
            inv_value_total = sum(v for _, v in excel_investments)
            all_inv_for_principal = db.query(InvestmentStatus).all()
            inv_principal_total: "float | None" = None
            principal_sum = sum(
                float(inv.principal)
                for inv in all_inv_for_principal
                if inv.principal is not None and float(inv.principal) > 0
            )
            if principal_sum > 0:
                inv_principal_total = principal_sum

            # 가장 마지막 월(현재 파일 기준)의 MonthlySummary에 저장
            if month_labels:
                last_label = month_labels[-1]
                try:
                    last_year, last_month = int(last_label[:4]), int(last_label[5:7])
                    latest_ms = db.query(MonthlySummary).filter(
                        MonthlySummary.year == last_year,
                        MonthlySummary.month == last_month,
                    ).first()
                    if latest_ms:
                        latest_ms.investment_value = inv_value_total
                        if inv_principal_total is not None:
                            latest_ms.investment_principal = inv_principal_total
                        db.commit()
                except (ValueError, IndexError):
                    pass

        # 5. 재무현황 (3.재무현황 섹션, rows 39~119)
        # col B(index 0)=카테고리, col C(index 1)=상품명, col E(index 3)=금액(자산), col I(index 7)=금액(부채)
        ASSET_CATEGORIES = {
            '자유입출금 자산', '신탁 자산', '현금 자산', '저축성 자산',
            '전자금융 자산', '투자성 자산', '부동산', '동산',
            '기타 실물 자산', '보험 자산', '연금 자산',
        }

        snap_data: dict = {}
        cur_cat: "str | None" = None
        total_assets_v: "float | None" = None
        total_liab_v: "float | None" = None
        net_assets_v: "float | None" = None

        for ridx in range(39, 120):
            r = rows.get(ridx)
            if r is None:
                continue
            label = r[0]          # col B
            product = r[1] if len(r) > 1 else None   # col C
            amt_e = r[3] if len(r) > 3 else None      # col E (자산 금액)
            amt_i = r[7] if len(r) > 7 else None      # col I (부채 금액)

            # 총자산/총부채 행
            if label == '총자산':
                if amt_e is not None:
                    total_assets_v = float(amt_e)
                if amt_i is not None:
                    total_liab_v = float(amt_i)
                continue

            # 순자산 값 행 (col B가 숫자인 경우)
            if label is not None and not isinstance(label, str):
                try:
                    net_assets_v = float(label)
                except (TypeError, ValueError):
                    pass
                continue

            # 자산 카테고리 헤더
            if isinstance(label, str) and label in ASSET_CATEGORIES:
                cur_cat = label
                if cur_cat not in snap_data:
                    snap_data[cur_cat] = []
                if product is not None and amt_e is not None:
                    snap_data[cur_cat].append({"name": str(product), "amount": float(amt_e)})
                continue

            # 카테고리 내 항목 (label=None, product 있음)
            if cur_cat and label is None and product is not None and amt_e is not None:
                snap_data[cur_cat].append({"name": str(product), "amount": float(amt_e)})

        # 부채 항목 파싱 (col F(r[4])=카테고리, col G(r[5])=상품명, col I(r[7])=금액)
        liab_data: dict = {}
        cur_liab: "str | None" = None
        SKIP_LIAB_LABELS = {'부채', '항목', None}
        for ridx in range(39, 118):
            r = rows.get(ridx)
            if r is None:
                continue
            liab_label = r[4] if len(r) > 4 else None   # col F
            liab_product = r[5] if len(r) > 5 else None  # col G
            liab_amt = r[7] if len(r) > 7 else None      # col I

            if liab_label == '총부채':
                break
            if isinstance(liab_label, str) and liab_label not in SKIP_LIAB_LABELS:
                cur_liab = liab_label.strip()
                liab_data.setdefault(cur_liab, [])
                if liab_product is not None and liab_amt is not None:
                    liab_data[cur_liab].append({"name": str(liab_product), "amount": float(liab_amt)})
            elif cur_liab and liab_label is None and liab_product is not None and liab_amt is not None:
                liab_data[cur_liab].append({"name": str(liab_product), "amount": float(liab_amt)})

        if liab_data:
            snap_data['_liabilities'] = liab_data

        # 의미 없는 빈 자산 카테고리 제거 (부채 키는 유지)
        snap_data = {
            k: v for k, v in snap_data.items()
            if k == '_liabilities' or v
        }

        # 총자산이 파싱되지 않은 경우(0 또는 None) snapshot_data 합계로 대체
        if not total_assets_v:
            total_assets_v = sum(
                item["amount"]
                for k, items in snap_data.items()
                if k != '_liabilities' and isinstance(items, list)
                for item in items
                if isinstance(item.get("amount"), (int, float))
            )
        # 총부채가 파싱되지 않은 경우 _liabilities 합계로 대체
        if not total_liab_v and '_liabilities' in snap_data:
            total_liab_v = sum(
                item["amount"]
                for items in snap_data['_liabilities'].values()
                if isinstance(items, list)
                for item in items
                if isinstance(item.get("amount"), (int, float))
            )
        if total_assets_v is not None and total_liab_v is not None and not net_assets_v:
            net_assets_v = total_assets_v - total_liab_v

        existing_snap = db.query(FinancialSnapshot).first()
        if existing_snap:
            existing_snap.total_assets = total_assets_v
            existing_snap.total_liabilities = total_liab_v
            existing_snap.net_assets = net_assets_v
            existing_snap.snapshot_data = snap_data
            result["financial_snapshot"]["updated"] = 1
        else:
            db.add(FinancialSnapshot(
                total_assets=total_assets_v,
                total_liabilities=total_liab_v,
                net_assets=net_assets_v,
                snapshot_data=snap_data,
            ))
            result["financial_snapshot"]["inserted"] = 1
        db.commit()

    # ── 가계부 내역 시트 ──────────────────────────────────────────
    if "가계부 내역" in wb.sheetnames:
        ws_ledger = wb["가계부 내역"]
        parsed_rows: list[dict] = []
        for row in ws_ledger.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            try:
                raw_date, raw_time, tx_type, category, subcategory, description, raw_amount, currency, payment_method, memo = row
            except ValueError:
                continue

            if isinstance(raw_date, datetime):
                tx_date = raw_date
            elif isinstance(raw_date, date):
                tx_date = datetime(raw_date.year, raw_date.month, raw_date.day)
            elif isinstance(raw_date, str) and raw_date.strip():
                try:
                    tx_date = datetime.fromisoformat(raw_date.strip())
                except ValueError:
                    continue
            else:
                continue

            amount_val: "float | None" = None
            if raw_amount is not None:
                try:
                    amount_val = float(str(raw_amount).replace(",", ""))
                except (ValueError, TypeError):
                    pass

            parsed_rows.append({
                "transaction_date": tx_date,
                "transaction_time": str(raw_time) if raw_time is not None else None,
                "transaction_type": str(tx_type) if tx_type else None,
                "category": str(category) if category else None,
                "subcategory": str(subcategory) if subcategory else None,
                "description": str(description) if description else None,
                "amount": amount_val,
                "currency": str(currency) if currency else None,
                "payment_method": str(payment_method) if payment_method else None,
                "memo": str(memo) if memo else None,
            })

        if parsed_rows:
            all_dates_l = [r["transaction_date"] for r in parsed_rows]
            min_dt, max_dt = min(all_dates_l), max(all_dates_l)
            existing_txns = (
                db.query(LedgerTransaction)
                .filter(
                    LedgerTransaction.transaction_date >= min_dt,
                    LedgerTransaction.transaction_date <= max_dt,
                )
                .all()
            )

            def _dedup_key_bs(tx_date, tx_time, desc, amt):
                return (
                    tx_date.strftime("%Y-%m-%d") if tx_date else None,
                    str(tx_time) if tx_time else None,
                    str(desc) if desc else None,
                    float(amt) if amt is not None else None,
                )

            existing_keys_bs = {
                _dedup_key_bs(e.transaction_date, e.transaction_time, e.description, e.amount)
                for e in existing_txns
            }

            for row in parsed_rows:
                k = _dedup_key_bs(row["transaction_date"], row["transaction_time"], row["description"], row["amount"])
                if k in existing_keys_bs:
                    result["ledger"]["skipped"] += 1
                else:
                    db.add(LedgerTransaction(**row))
                    existing_keys_bs.add(k)  # 같은 배치 내 중복 삽입 방지
                    result["ledger"]["inserted"] += 1

            if result["ledger"]["inserted"] > 0:
                db.commit()

    # 업로드 이력 저장
    history = UploadHistory(
        filename=file.filename,
        file_size=file_size,
        result_json=result,
    )
    db.add(history)
    db.commit()

    return result


@router.post("/ledger-transactions/import-excel")
async def import_ledger_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """가계부 내역 Excel 파일(뱅크샐러드 형식)을 읽어 신규 거래만 DB에 추가합니다."""
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="xlsx 또는 xls 파일만 지원합니다.")

    file_bytes = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 파일을 읽을 수 없습니다.")

    if "가계부 내역" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail='"가계부 내역" 시트를 찾을 수 없습니다.')

    ws = wb["가계부 내역"]

    # ── 행 파싱 ──────────────────────────────────────────────
    parsed: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        try:
            raw_date, raw_time, tx_type, category, subcategory, description, raw_amount, currency, payment_method, memo = row
        except ValueError:
            continue

        # 날짜 → datetime
        if isinstance(raw_date, datetime):
            tx_date = raw_date
        elif isinstance(raw_date, date):
            tx_date = datetime(raw_date.year, raw_date.month, raw_date.day)
        elif isinstance(raw_date, str) and raw_date.strip():
            try:
                tx_date = datetime.fromisoformat(raw_date.strip())
            except ValueError:
                continue
        else:
            continue  # 날짜 없으면 스킵

        # 금액 정규화
        amount_val: float | None = None
        if raw_amount is not None:
            try:
                amount_val = float(str(raw_amount).replace(",", ""))
            except (ValueError, TypeError):
                pass

        parsed.append({
            "transaction_date": tx_date,
            "transaction_time": str(raw_time) if raw_time is not None else None,
            "transaction_type": str(tx_type) if tx_type else None,
            "category": str(category) if category else None,
            "subcategory": str(subcategory) if subcategory else None,
            "description": str(description) if description else None,
            "amount": amount_val,
            "currency": str(currency) if currency else None,
            "payment_method": str(payment_method) if payment_method else None,
            "memo": str(memo) if memo else None,
        })

    if not parsed:
        return {"inserted": 0, "skipped": 0}

    # ── 중복 제거 (date + time + description + amount) ───────
    all_dates = [r["transaction_date"] for r in parsed]
    min_date, max_date = min(all_dates), max(all_dates)

    existing = (
        db.query(LedgerTransaction)
        .filter(
            LedgerTransaction.transaction_date >= min_date,
            LedgerTransaction.transaction_date <= max_date,
        )
        .all()
    )

    def _dedup_key(tx_date, tx_time, desc, amt):
        return (
            tx_date.strftime("%Y-%m-%d") if tx_date else None,
            str(tx_time) if tx_time else None,
            str(desc) if desc else None,
            float(amt) if amt is not None else None,
        )

    existing_keys = {
        _dedup_key(e.transaction_date, e.transaction_time, e.description, e.amount)
        for e in existing
    }

    # ── INSERT 신규 거래 ──────────────────────────────────────
    inserted = 0
    skipped = 0
    for row in parsed:
        key = _dedup_key(row["transaction_date"], row["transaction_time"], row["description"], row["amount"])
        if key in existing_keys:
            skipped += 1
        else:
            db.add(LedgerTransaction(**row))
            inserted += 1

    if inserted > 0:
        db.commit()

    return {"inserted": inserted, "skipped": skipped}
