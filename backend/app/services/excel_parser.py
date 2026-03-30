import openpyxl
import xlrd
from typing import List, Dict, Any, Optional
import io
from datetime import datetime, date, time
import logging

logger = logging.getLogger(__name__)

def convert_datetime_to_string(value: Any) -> Any:
    """
    datetime, date, time 객체를 ISO 형식 문자열로 변환합니다.
    다른 타입의 값은 그대로 반환합니다.
    
    Args:
        value: 변환할 값
    
    Returns:
        datetime/date/time 객체는 ISO 형식 문자열, 그 외는 원본 값
    """
    if isinstance(value, datetime):
        return value.isoformat()
    elif isinstance(value, date):
        return value.isoformat()
    elif isinstance(value, time):
        return value.isoformat()
    return value

def parse_excel_file(file_content: bytes, filename: str) -> Dict[str, Any]:
    """
    엑셀 파일을 파싱하여 데이터를 반환합니다.
    
    Args:
        file_content: 파일 바이너리 내용
        filename: 파일명 (확장자로 형식 판단)
    
    Returns:
        {
            "sheets": [
                {
                    "name": "시트명",
                    "data": [[행1], [행2], ...],
                    "row_count": 행 수,
                    "column_count": 열 수
                }
            ]
        }
    """
    sheets_data = []
    
    if filename.endswith('.xlsx'):
        # data_only=True로 먼저 로드, formula 워크북은 None 셀 발견 시에만 lazy 로딩 (P2)
        logger.info(f"워크북 (data_only=True) 로딩 시작...")
        workbook_data_only = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        logger.info(f"워크북 (data_only=True) 로딩 완료")

        workbook_formula = None  # lazy: None 셀이 존재할 때만 로드

        def _get_formula_sheet(sname: str):
            nonlocal workbook_formula
            if workbook_formula is None:
                logger.info("수식 확인 필요 — formula 워크북 lazy 로딩 시작")
                workbook_formula = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)
                logger.info("formula 워크북 로딩 완료")
            return workbook_formula[sname]

        # 연속 빈 행이 이 수를 초과하면 시트 처리 종료 (Excel ghost row 방지)
        CONSECUTIVE_EMPTY_LIMIT = 50

        for sheet_name in workbook_data_only.sheetnames:
            sheet_data_only = workbook_data_only[sheet_name]
            data = []

            logger.info(
                f"시트 '{sheet_name}' 데이터 읽기 시작 "
                f"(최대 행: {sheet_data_only.max_row}, 최대 열: {sheet_data_only.max_column})"
            )
            consecutive_empty = 0
            for row_num, row_data_only in enumerate(
                sheet_data_only.iter_rows(values_only=True), 1
            ):
                if row_num % 1000 == 0:
                    logger.info(f"시트 '{sheet_name}' {row_num}행 처리 중...")

                # 행 전체가 비어있으면 formula 워크북 참조 없이 바로 skip
                if all(cell is None or cell == "" for cell in row_data_only):
                    consecutive_empty += 1
                    if consecutive_empty >= CONSECUTIVE_EMPTY_LIMIT:
                        logger.info(
                            f"시트 '{sheet_name}': 연속 빈 행 {CONSECUTIVE_EMPTY_LIMIT}개 도달, "
                            f"처리 종료 (마지막 처리 행: {row_num})"
                        )
                        break
                    continue
                consecutive_empty = 0

                row_data = []
                for col_idx, cell_value in enumerate(row_data_only, 1):
                    # 계산된 값이 None/빈 값이면 formula 워크북을 lazy 로드해 확인
                    if cell_value is None or cell_value == "":
                        formula_sheet = _get_formula_sheet(sheet_name)
                        cell_formula_obj = formula_sheet.cell(row=row_num, column=col_idx)
                        if cell_formula_obj.data_type == 'f':
                            formula = cell_formula_obj.value
                            if not isinstance(formula, str):
                                logger.debug(
                                    f"셀 {openpyxl.utils.get_column_letter(col_idx)}{row_num}: "
                                    f"복잡한 수식 타입 ({type(formula).__name__}), 0으로 설정"
                                )
                                cell_value = 0
                            else:
                                logger.warning(
                                    f"셀 {openpyxl.utils.get_column_letter(col_idx)}{row_num}: "
                                    f"수식이지만 계산값 없음 ({formula})"
                                )
                                calculated = _try_calculate_simple_formula(
                                    formula, sheet_data_only, row_num, col_idx
                                )
                                cell_value = calculated if calculated is not None else 0

                    cell_result = convert_datetime_to_string(cell_value) if cell_value is not None else ""
                    row_data.append(cell_result)

                if any(cell != "" for cell in row_data):
                    data.append(row_data)
            
            if data:
                # 최대 열 수 계산
                max_cols = max(len(row) for row in data) if data else 0
                sheets_data.append({
                    "name": sheet_name,
                    "data": data,
                    "row_count": len(data),
                    "column_count": max_cols
                })
    
    elif filename.endswith('.xls'):
        # xlrd로 xls 파일 파싱
        workbook = xlrd.open_workbook(file_contents=file_content)
        
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            data = []
            
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    # xlrd는 날짜를 float로 반환하므로, 날짜 타입인 경우 datetime으로 변환 후 문자열로 변환
                    if sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                        # xlrd의 날짜 변환
                        date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                        if date_tuple:
                            cell_value = datetime(*date_tuple[:6])
                    row_data.append(convert_datetime_to_string(cell_value))
                
                if any(cell != "" for cell in row_data):  # 빈 행 제외
                    data.append(row_data)
            
            if data:
                max_cols = max(len(row) for row in data) if data else 0
                sheets_data.append({
                    "name": sheet_name,
                    "data": data,
                    "row_count": len(data),
                    "column_count": max_cols
                })
    else:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {filename}")
    
    return {
        "sheets": sheets_data,
        "sheet_count": len(sheets_data)
    }

def _try_calculate_simple_formula(
    formula: Any, 
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    col_idx: int
) -> Optional[Any]:
    """
    간단한 수식을 계산하려고 시도합니다.
    현재는 SUM, AVERAGE, COUNT, MIN, MAX 등의 기본 함수만 지원합니다.
    
    Args:
        formula: 수식 문자열 또는 수식 객체 (예: "=SUM(A1:A5)")
        sheet: 워크시트 객체
        row_idx: 현재 행 인덱스 (1-based)
        col_idx: 현재 열 인덱스 (1-based)
    
    Returns:
        계산된 값 또는 None (계산할 수 없는 경우)
    """
    # formula가 문자열이 아니면 처리 불가 (ArrayFormula 등)
    if not isinstance(formula, str):
        logger.debug(f"수식이 문자열이 아닙니다. 타입: {type(formula)}")
        return None
    
    if not formula or not formula.startswith('='):
        return None
    
    try:
        # SUM 함수 처리
        if formula.upper().startswith('=SUM('):
            # SUM 범위 추출 및 계산
            # 괄호 안의 내용 추출
            inner = formula[5:-1].strip()  # '=SUM(' 와 ')' 제거
            
            # 여러 인자가 있을 수 있으므로 쉼표로 분리
            # 하지만 범위 내에도 쉼표가 있을 수 있으므로 간단하게 첫 번째 인자만 처리
            # 실제로는 더 복잡한 파싱이 필요하지만, 기본적인 경우만 처리
            args = []
            current_arg = ""
            paren_depth = 0
            
            for char in inner:
                if char == '(':
                    paren_depth += 1
                    current_arg += char
                elif char == ')':
                    paren_depth -= 1
                    current_arg += char
                elif char == ',' and paren_depth == 0:
                    if current_arg.strip():
                        args.append(current_arg.strip())
                    current_arg = ""
                else:
                    current_arg += char
            
            if current_arg.strip():
                args.append(current_arg.strip())
            
            # 모든 인자의 합계 계산
            total = 0.0
            for arg in args:
                arg_result = _calculate_sum_range(arg, sheet)
                if arg_result is not None:
                    total += arg_result
            
            return total if total != 0.0 else None
        
        # 추가 함수들을 여기에 구현할 수 있습니다.
        # 현재는 SUM만 지원
        
    except Exception as e:
        logger.debug(f"수식 계산 중 오류 발생: {formula}, 오류: {e}")
        return None
    
    return None

def _calculate_sum_range(range_str: str, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[float]:
    """
    SUM 범위를 계산합니다.
    
    Args:
        range_str: 범위 문자열 (예: "A1:A5", "A1:B2")
        sheet: 워크시트 객체
    
    Returns:
        합계 값 또는 None
    """
    try:
        # 범위 파싱 (예: "A1:A5" -> (1, 1) to (5, 1))
        if ':' in range_str:
            start_cell, end_cell = range_str.split(':')
            start_col = openpyxl.utils.column_index_from_string(
                ''.join(c for c in start_cell if c.isalpha())
            )
            start_row = int(''.join(c for c in start_cell if c.isdigit()))
            end_col = openpyxl.utils.column_index_from_string(
                ''.join(c for c in end_cell if c.isalpha())
            )
            end_row = int(''.join(c for c in end_cell if c.isdigit()))
            
            total = 0.0
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)):
                        total += float(cell_value)
            return total
        else:
            # 단일 셀 참조
            col = openpyxl.utils.column_index_from_string(
                ''.join(c for c in range_str if c.isalpha())
            )
            row = int(''.join(c for c in range_str if c.isdigit()))
            cell_value = sheet.cell(row=row, column=col).value
            if isinstance(cell_value, (int, float)):
                return float(cell_value)
    except Exception as e:
        logger.debug(f"SUM 범위 계산 중 오류 발생: {range_str}, 오류: {e}")
        return None
    
    return None

